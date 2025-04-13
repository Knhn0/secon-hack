import http

from fastapi import UploadFile
from openpyxl.styles import PatternFill, Alignment, Font
from typing import List
from openpyxl import load_workbook

from services.ai import AIService
from services.storage import StorageService
from utils.unitofwork import IUnitOfWork


class ReportService:
    async def get_report(self, uow: IUnitOfWork, _id: int):
        return await uow.reports.find_one(id=_id)

    async def delete_report(self, uow: IUnitOfWork, _id: int):
        report_id = await uow.reports.delete_one(_id)
        await uow.commit()
        return report_id

    async def get_all(self, uow: IUnitOfWork):
        return await uow.reports.find_all()

    async def update_excel(self, filename, names):
        wb = load_workbook(filename)
        ws = wb.active

        ws.cell(row=8, column=45).value = "Ссылка на фото"
        ws.cell(row=8, column=45).font = Font(bold=True)
        ws.cell(row=8, column=45).alignment = Alignment(horizontal="center", vertical="center")
        wb.save(filename)

        fill = PatternFill(
            start_color="CCFF99",  # салатовый
            end_color="CCFF99",
            fill_type="solid"
        )

        for name in names:
            link = await StorageService().get_link(name)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == name:
                        cell.fill = fill
                        cell(colomn=45).value = link
                        wb.save(filename)
                        break

    async def post_report(files: List[UploadFile]):
        jpgNames = List[str]
        for file in files:
            if file.filename.endswith((".xls", ".xlsx")):
                continue
            recognized_text = await AIService().handle_file(file)
            if recognized_text == "Номер не найден":
                file.filename = f"unrecognized_number_{file.filename}"
            else:
                jpgNames.append(file.filename)
            await StorageService().upload_file(file)

        for file in files:
            if file.filename.endswith('.xlsx'):
                await ReportService.update_excel(file.filename, jpgNames)
                await StorageService().upload_file(file)
        return http.HTTPStatus.OK
